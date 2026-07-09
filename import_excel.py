"""Aylık Ödeme Takip Dosyası'nı (Excel) veritabanına aktarır.

Çalıştırma: python import_excel.py
İdempotenttir: ikinci çalıştırmada mevcut kayıtlar atlanır.

Sayfalar:
- DOĞALGAZ AVANS      → daireler (blok, no, niteliği, m²)
- DEMIRBAS_TAKIP      → kat malikleri + demirbaş borç/ödemeleri (Temmuz–Eylül 2026)
- AIDAT_DOGALGAZ_TAKIP→ aidat tanımları + borç/ödemeleri (2024)
- DOGALGAZ_DATA       → daire bazlı doğalgaz borçları (tutar > 0 olanlar)

Not: takip sayfalarındaki BLOK kolonu güvenilmez (C-Blok dükkanlar "B" yazılmış);
blok/daire her zaman etiket kolonundan (ör. "C-BLOK DÜKKAN 01") çözülür.

Not: Excel başlıklarındaki 2024 tarihleri bayattır — site Haziran 2026'da
açıldı; aidat/doğalgaz dönemleri YEAR_OVERRIDE ile 2026'ya çevrilir.
"""

import re
import sys
import unicodedata
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from sqlalchemy import select

from app import models
from app.auth import hash_password
from app.database import Base, SessionLocal, engine

EXCEL_PATH = "Aylık Ödeme Takip Dosyası - 2.xlsx"
SITE_NAME = "Troaspark Sitesi"
OWNER_PASSWORD = "demo1234"
EMAIL_DOMAIN = "troaspark.local"
DEMIRBAS_YEAR = 2026  # Temmuz–Eylül kolonları bu yıla aittir
DEMIRBAS_MONTHS = {"Temmuz": 7, "Ağustos": 8, "Eylül": 9}
# Excel başlıkları 2024 yazsa da gerçek dönemler 2026 (site açılışı: Haziran 2026)
YEAR_OVERRIDE = {2024: 2026}
SITE_OPENED = date(2026, 6, 1)

_TR_MAP = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosucgiosu")


def slug_email(full_name: str) -> str:
    """"Sinan Ünsal" → "sinan.unsal@troaspark.local"."""
    ascii_name = unicodedata.normalize("NFKD", full_name.translate(_TR_MAP))
    ascii_name = ascii_name.encode("ascii", "ignore").decode()
    parts = re.findall(r"[a-z0-9]+", ascii_name.lower())
    return ".".join(parts) + "@" + EMAIL_DOMAIN


def parse_unit_label(label: str) -> tuple[str, int] | None:
    """Etiketten (blok harfi, daire no) çözer.

    "A-BLOK DAİRE 02" → ("A", 2); "A-BLOK KREŞ" → ("A", 1);
    "B-BLOK KAPICI D." → ("B", 0); "C-BLOK DÜKKAN 01" → ("C", 1)
    """
    if not label:
        return None
    text = str(label).strip().upper()
    m = re.match(r"([A-C])\s*-?\s*BLOK", text)
    if not m:
        return None
    block = m.group(1)
    if "KREŞ" in text:
        return block, 1
    if "KAPICI" in text:
        return block, 0
    digits = re.search(r"(\d+)\s*$", text)
    if not digits:
        return None
    return block, int(digits.group(1))


def parse_gas_unit(block_name: str, unit_name: str) -> tuple[str, int] | None:
    """DOGALGAZ_DATA satırındaki ("A-BLOK", "DAİRE 02") çiftini çözer."""
    if not block_name or not unit_name:
        return None
    return parse_unit_label(f"{block_name} {unit_name}")


def month_column_pairs(header_row: list, max_col: int) -> list[tuple[str, int, int]]:
    """Başlık satırındaki tarih/ay çapalarından (dönem, tahakkuk_idx, ödeme_idx) üretir.

    Her çapanın kolon çifti, bir sonraki çapadan hemen önceki iki kolondur;
    böylece 2024-01'deki fazladan Tahakkuk kolonu doğal olarak atlanır.
    """
    anchors: list[tuple[int, str]] = []
    for idx, value in enumerate(header_row):
        if isinstance(value, datetime):
            year = YEAR_OVERRIDE.get(value.year, value.year)
            anchors.append((idx, f"{year:04d}-{value.month:02d}"))
        elif isinstance(value, str) and value.strip() in DEMIRBAS_MONTHS:
            month = DEMIRBAS_MONTHS[value.strip()]
            anchors.append((idx, f"{DEMIRBAS_YEAR:04d}-{month:02d}"))
        elif isinstance(value, str) and value.strip() == "TOPLAM":
            anchors.append((idx, "TOPLAM"))
    pairs = []
    for i, (idx, period) in enumerate(anchors):
        if period == "TOPLAM":
            continue
        end = anchors[i + 1][0] if i + 1 < len(anchors) else max_col
        pairs.append((period, end - 2, end - 1))
    return pairs


def parse_money(value) -> Decimal | None:
    """Hücre değerini Decimal'e çevirir; boş/0/sayı olmayan için None."""
    if value is None:
        return None
    try:
        amount = Decimal(str(value).strip().replace(",", "."))
    except InvalidOperation:
        return None
    return amount if amount > 0 else None


def month_end(period: str) -> date:
    year, month = int(period[:4]), int(period[5:7])
    nxt = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    return nxt - timedelta(days=1)


def main() -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl gerekli: .venv/bin/pip install -r requirements.txt")

    Base.metadata.create_all(bind=engine)
    wb = load_workbook(EXCEL_PATH, data_only=True)
    db = SessionLocal()
    stats = {"apartment": 0, "user": 0, "occupancy": 0, "debt": 0, "payment": 0, "dues": 0}

    try:
        # --- Site + bloklar ---
        site = db.scalar(select(models.Site).where(models.Site.name == SITE_NAME))
        if site is None:
            site = models.Site(name=SITE_NAME)
            db.add(site)
            db.flush()
        blocks: dict[str, models.Block] = {}
        for letter in ("A", "B", "C"):
            block = db.scalar(
                select(models.Block).where(
                    models.Block.site_id == site.id, models.Block.name == f"{letter} Blok"
                )
            )
            if block is None:
                block = models.Block(site_id=site.id, name=f"{letter} Blok")
                db.add(block)
                db.flush()
            blocks[letter] = block

        apartments: dict[tuple[str, int], models.Apartment] = {}
        for apartment in db.scalars(
            select(models.Apartment)
            .join(models.Block)
            .where(models.Block.site_id == site.id)
        ):
            letter = apartment.block.name[0]
            apartments[(letter, int(apartment.number))] = apartment

        def get_or_create_apartment(
            letter: str, number: int, unit_type=None, area=None
        ) -> models.Apartment:
            key = (letter, number)
            if key not in apartments:
                apartment = models.Apartment(
                    block_id=blocks[letter].id,
                    floor_no=0,
                    number=str(number),
                    unit_type=unit_type,
                    area_m2=float(area) if area else None,
                )
                db.add(apartment)
                db.flush()
                apartments[key] = apartment
                stats["apartment"] += 1
            return apartments[key]

        # --- Daireler (DOĞALGAZ AVANS: tam liste, niteliği + m²) ---
        for row in wb["DOĞALGAZ AVANS"].iter_rows(min_row=2, values_only=True):
            letter, number = str(row[1] or "").strip(), row[2]
            if letter not in blocks or not isinstance(number, (int, float)):
                continue
            get_or_create_apartment(
                letter,
                int(number),
                unit_type=str(row[4]).strip() if row[4] else None,
                area=row[5],
            )

        # --- Kat malikleri + kiracılar (DEMIRBAS_TAKIP) ---
        password_hash = hash_password(OWNER_PASSWORD)
        users_by_name: dict[str, models.User] = {}
        for user in db.scalars(select(models.User)):
            users_by_name[user.full_name.strip().casefold()] = user

        def get_or_create_user(full_name: str, phone=None) -> models.User:
            key = full_name.strip().casefold()
            if key not in users_by_name:
                email = slug_email(full_name)
                user = models.User(
                    email=email,
                    full_name=full_name.strip(),
                    phone=str(phone).strip() if phone else None,
                    role=models.ROLE_RESIDENT,
                    password_hash=password_hash,
                )
                db.add(user)
                db.flush()
                users_by_name[key] = user
                stats["user"] += 1
            return users_by_name[key]

        def ensure_occupancy(apartment: models.Apartment, user: models.User, occ_type: str):
            exists = db.scalar(
                select(models.Occupancy).where(
                    models.Occupancy.apartment_id == apartment.id,
                    models.Occupancy.user_id == user.id,
                    models.Occupancy.type == occ_type,
                    models.Occupancy.end_date.is_(None),
                )
            )
            if exists is None:
                db.add(
                    models.Occupancy(
                        apartment_id=apartment.id, user_id=user.id, type=occ_type,
                        start_date=SITE_OPENED,
                    )
                )
                stats["occupancy"] += 1

        existing_debts: set[tuple[int, str, str]] = {
            (d.apartment_id, d.category, d.description)
            for d in db.scalars(
                select(models.Debt)
                .join(models.Apartment)
                .join(models.Block)
                .where(models.Block.site_id == site.id)
            )
        }

        def add_debt(
            apartment: models.Apartment,
            category: str,
            description: str,
            amount: Decimal,
            due: date,
            paid: Decimal | None,
            dues_id: int | None = None,
        ) -> None:
            key = (apartment.id, category, description)
            if key in existing_debts:
                return
            debt = models.Debt(
                apartment_id=apartment.id,
                dues_id=dues_id,
                description=description,
                amount=amount,
                due_date=due,
                category=category,
                bill_to_owner=category in models.OWNER_BILLED_CATEGORIES,
            )
            db.add(debt)
            db.flush()
            existing_debts.add(key)
            stats["debt"] += 1
            if paid:
                db.add(
                    models.Payment(
                        debt_id=debt.id, amount=paid, paid_at=due, method="transfer"
                    )
                )
                stats["payment"] += 1
                if paid >= amount:
                    debt.status = models.DEBT_PAID
                elif paid > 0:
                    debt.status = models.DEBT_PARTIAL

        # --- Demirbaş: malik + borç/ödeme ---
        ws = wb["DEMIRBAS_TAKIP"]
        header = [c.value for c in ws[1]]
        pairs = month_column_pairs(header, ws.max_column)
        for row in ws.iter_rows(min_row=3, values_only=True):
            unit = parse_unit_label(row[1])
            if unit is None:
                continue
            apartment = get_or_create_apartment(*unit)
            owner_name = str(row[6] or "").strip()
            if owner_name and owner_name.casefold() != "kapıcı dairesi":
                owner = get_or_create_user(owner_name, phone=row[7])
                ensure_occupancy(apartment, owner, models.OCC_OWNER)
                tenant_name = str(row[8] or "").strip()
                if tenant_name:
                    tenant = get_or_create_user(tenant_name, phone=row[9])
                    ensure_occupancy(apartment, tenant, models.OCC_TENANT)
            for period, tah_idx, ode_idx in pairs:
                amount = parse_money(row[tah_idx])
                if amount is None:
                    continue
                add_debt(
                    apartment,
                    models.DEBT_CAT_DEMIRBAS,
                    f"{period} demirbaş taksidi",
                    amount,
                    month_end(period),
                    parse_money(row[ode_idx]),
                )

        # --- Aidat: dönem tanımları + borç/ödeme ---
        ws = wb["AIDAT_DOGALGAZ_TAKIP"]
        header = [c.value for c in ws[1]]
        pairs = month_column_pairs(header, ws.max_column)
        dues_by_period: dict[str, models.DuesDefinition] = {}
        for dues in db.scalars(
            select(models.DuesDefinition).where(
                models.DuesDefinition.site_id == site.id,
                models.DuesDefinition.is_surcharge.is_(False),
            )
        ):
            dues_by_period[dues.period] = dues

        def get_or_create_dues(period: str, amount: Decimal) -> models.DuesDefinition:
            if period not in dues_by_period:
                dues = models.DuesDefinition(
                    site_id=site.id,
                    block_id=None,
                    period=period,
                    amount=amount,
                    due_date=month_end(period),
                    description=f"{period} aidatı",
                )
                db.add(dues)
                db.flush()
                dues_by_period[period] = dues
                stats["dues"] += 1
            return dues_by_period[period]

        for row in ws.iter_rows(min_row=3, values_only=True):
            unit = parse_unit_label(row[1])
            if unit is None:
                continue
            apartment = get_or_create_apartment(*unit)
            for period, tah_idx, ode_idx in pairs:
                amount = parse_money(row[tah_idx])
                if amount is None:
                    continue
                dues = get_or_create_dues(period, amount)
                add_debt(
                    apartment,
                    models.DEBT_CAT_AIDAT,
                    f"{period} aidatı",
                    amount,
                    dues.due_date,
                    parse_money(row[ode_idx]),
                    dues_id=dues.id,
                )

        # --- Doğalgaz abonelik durumu (YAPILDI / YAPILMADI '*' işaretli) ---
        for row in wb["DOĞALGAZ ABONELİK"].iter_rows(min_row=2, values_only=True):
            unit = parse_unit_label(row[1])
            if unit is None:
                continue
            apartment = get_or_create_apartment(*unit)
            done = str(row[8] or "").strip()
            not_done = str(row[7] or "").strip()
            if done:
                apartment.gas_subscribed = True
            elif not_done:
                apartment.gas_subscribed = False

        # --- Doğalgaz (DOGALGAZ_DATA: tutar > 0 olan dönem hücreleri) ---
        ws = wb["DOGALGAZ_DATA"]
        rows = list(ws.iter_rows(values_only=True))
        gas_header = rows[4]  # ['Sütun1', 'BLOK', 'DAİRE', 'MT²', '1.01.2024', ...]
        gas_periods: dict[int, str] = {}
        for idx, value in enumerate(gas_header[4:], start=4):
            m = re.match(r"\d{1,2}\.(\d{2})\.(\d{4})", str(value or ""))
            if m:
                year = YEAR_OVERRIDE.get(int(m.group(2)), int(m.group(2)))
                gas_periods[idx] = f"{year:04d}-{m.group(1)}"
        for row in rows[5:]:
            unit = parse_gas_unit(row[1], row[2])
            if unit is None:
                continue
            apartment = get_or_create_apartment(*unit)
            for idx, period in gas_periods.items():
                amount = parse_money(row[idx])
                if amount is None:
                    continue
                add_debt(
                    apartment,
                    models.DEBT_CAT_DOGALGAZ,
                    f"{period} doğalgaz payı",
                    amount,
                    month_end(period),
                    None,
                )

        db.commit()
        print(
            f"İçe aktarma tamamlandı ({SITE_NAME}): "
            f"{stats['apartment']} daire, {stats['user']} kişi, {stats['occupancy']} malik/kiracı ataması, "
            f"{stats['dues']} aidat tanımı, {stats['debt']} borç, {stats['payment']} ödeme oluşturuldu."
        )
        if stats["user"]:
            print(f"Malik hesapları: ad.soyad@{EMAIL_DOMAIN} / parola: {OWNER_PASSWORD}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
